import openpyxl, re, shutil, time, os, sys, tkinter as tk
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from datetime import datetime
from tkinter import messagebox, ttk
from tkcalendar import DateEntry
import threading
import tempfile
import logging
from logging.handlers import RotatingFileHandler
import pythoncom
import win32com.client as win32

BASE_DIR = os.path.dirname(sys.executable if getattr(sys, 'frozen', False) else os.path.abspath(__file__))
CARPETA, ARCHIVO = "reportes", "Reporte IR Tornos.xlsx"
RUTA_ENTRADA = os.path.join(BASE_DIR, CARPETA, ARCHIVO)
MESES = dict(zip(
    ["January","February","March","April","May","June","July","August","September","October","November","December"],
    ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
))
MESES_NUM = {
    "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4,
    "Mayo": 5, "Junio": 6, "Julio": 7, "Agosto": 8,
    "Septiembre": 9, "Octubre": 10, "Noviembre": 11, "Diciembre": 12
}
BORDER = Border(*(Side(style='thin'),)*4)
ALIGN_R = Alignment(horizontal='right')
FILL_AMARILLO = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def configurar_logging():
    # Configuracion de un sistema de logging
    posibles_rutas = [
        os.path.join(BASE_DIR, CARPETA, "log_tornos.log"),
        os.path.join(tempfile.gettempdir(), "log_tornos.log")
    ]
    logger = logging.getLogger('TornosLogger')
    logger.setLevel(logging.INFO)
    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    for ruta in posibles_rutas:
        try:
            os.makedirs(os.path.dirname(ruta), exist_ok=True)
            handler = RotatingFileHandler(
                ruta,
                maxBytes=5*1024*1024,
                backupCount=3,
                encoding='utf-8'
            )
            handler.setFormatter(formatter)
            logger.addHandler(handler)
            return logger
        except Exception as e:
            escribir_log(f" \n No se pudo configurar log en {ruta}: {e}", file=sys.stderr)
    # Si fallan todas las rutas, crear logger de consola
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)
    logger.warning("No se pudo crear archivo de log. Usando consola.")
    return logger
logger = configurar_logging()

def escribir_log(mensaje, nivel="info"):
    # Funcion para escribir en el log
    try:
        if nivel.lower() == "info":
            logger.info(mensaje)
        elif nivel.lower() == "warning":
            logger.warning(mensaje)
        elif nivel.lower() == "error":
            logger.error(mensaje)
        else:
            logger.debug(mensaje)
    except Exception as e:
        escribir_log(f"Error al escribir en log: {e}", file=sys.stderr)

def pedir_fecha(callback):
    ventana = tk.Toplevel()
    ventana.title("Fecha del reporte")
    ventana.geometry("300x200")
    ventana.resizable(False, False)
    tk.Label(ventana, text="Selecciona la fecha:").pack(pady=10)
    ent_fecha = DateEntry(ventana, date_pattern='dd/MM/yyyy')
    ent_fecha.pack(pady=10); ent_fecha.set_date(datetime.now())
    def confirmar():
        f = ent_fecha.get_date()
        callback(MESES[f.strftime("%B")], f.day, f.year)
        ventana.destroy()
    tk.Button(ventana, text="Aceptar", command=confirmar).pack(pady=10)
    ventana.grab_set()

def buscar_archivos_torno(fecha):
    """
    Busca los archivos de reporte para los tornos 3011 y 3012 en la fecha especificada.
    Retorna una tupla con las rutas (torno1_path, torno2_path) o None si no se encuentran.
    """
    fecha_str = fecha.strftime("%d-%m-%Y")
    patron_torno1 = f"Reporte_{fecha_str}_3011.txt"
    patron_torno2 = f"Reporte_{fecha_str}_3012.txt"
    posibles_rutas = [
        os.path.join(BASE_DIR, "Reportes_Tornos", "datos"),
        os.path.join(BASE_DIR, "..", "Reportes_Tornos", "datos"),
        os.path.join(BASE_DIR, "..", "..", "Reportes_Tornos", "datos"),
        os.path.join(os.path.dirname(BASE_DIR), "Reportes_Tornos", "datos")
    ]
    archivos_encontrados = {"3011": None, "3012": None}
    
    for ruta in posibles_rutas:
        try:
            if not os.path.isdir(ruta):
                escribir_log(f"Directorio no encontrado: {ruta}", nivel="debug")
                continue
            # Listar archivos en el directorio
            for archivo in os.listdir(ruta):
                if archivo == patron_torno1:
                    archivos_encontrados["3011"] = os.path.join(ruta, archivo)
                elif archivo == patron_torno2:
                    archivos_encontrados["3012"] = os.path.join(ruta, archivo)
            # Si se encuentran los archivos salir del bucle
            if all(archivos_encontrados.values()):
                break
                
        except Exception as e:
            escribir_log(f"Error buscando archivos en {ruta}: {str(e)}", nivel="warning")
    
    # Si no se encuentran las rutas
    if not all(archivos_encontrados.values()):
        escribir_log("No se encontraron archivos en rutas estándar, intentando búsqueda flexible...", nivel="debug")
        for ruta in posibles_rutas:
            try:
                if not os.path.isdir(ruta):
                    continue

                for archivo in os.listdir(ruta):
                    if fecha_str in archivo and archivo.endswith(".txt"):
                        if "3011" in archivo:
                            archivos_encontrados["3011"] = os.path.join(ruta, archivo)
                        elif "3012" in archivo:
                            archivos_encontrados["3012"] = os.path.join(ruta, archivo)
                            
                if all(archivos_encontrados.values()):
                    break
                    
            except Exception as e:
                escribir_log(f"Error en búsqueda flexible en {ruta}: {str(e)}", nivel="warning")
    
    # Registrar resultados de la búsqueda
    escribir_log(f"Resultados de búsqueda - Torno1: encontrado, Torno2: encontrado")
    return archivos_encontrados["3011"], archivos_encontrados["3012"]

def leer_archivo_torno(ruta_archivo):
    # Leer el contenido del archivo de reporte de torno

    if not ruta_archivo or not os.path.exists(ruta_archivo):
        escribir_log(f"Archivo no encontrado: {ruta_archivo}", nivel="warning")
        return None

    try:
        with open(ruta_archivo, 'r', encoding='utf-8') as f:
            contenido = f.read()

        # Normalizar saltos de línea y limpiar espacios
        contenido = contenido.replace('\r\n', '\n').replace('\r', '\n').strip()

        # Verificar que el contenido tenga datos válidos
        if not contenido:
            escribir_log(f"Archivo vacío: {ruta_archivo}", nivel="warning")
            return None

        # Verificar si contiene una línea con "RADIATA"
        if "RADIATA" not in contenido:
            escribir_log(f"Archivo no contiene datos RADIATA: {ruta_archivo}", nivel="warning")
            return None

        return contenido

    except Exception as e:
        escribir_log(f"Error leyendo archivo {ruta_archivo}: {str(e)}", nivel="error")
        return None

def iniciar(texto, torno, mes, dia, anio):
    mostrar_carga()
    threading.Thread(target=lambda: ejecutar(texto, torno, mes, dia, anio), daemon=True).start()

def mostrar_carga():
    # Muestra la ventana de carga
    global ventana_carga, barra
    if 'ventana_carga' not in globals() or not ventana_carga.winfo_exists():
        ventana_carga = tk.Toplevel()
        ventana_carga.title("Procesando datos...")
        ventana_carga.geometry("400x120")
        ventana_carga.resizable(False, False)
        ventana_carga.protocol("WM_DELETE_WINDOW", lambda: None)
        tk.Label(ventana_carga, 
                text="Procesando datos...", 
                font=("Arial", 12)).pack(pady=10)
        barra = ttk.Progressbar(ventana_carga, mode='determinate', maximum=100)
        barra.pack(fill='x', padx=20, pady=5)
        # Mostrar el torno actual
        global lbl_estado
        lbl_estado = tk.Label(ventana_carga, text="", font=("Arial", 10))
        lbl_estado.pack()
        ventana_carga.grab_set()

    barra['value'] = 0
    ventana_carga.deiconify()

def cerrar_carga():
    if ventana_carga: ventana_carga.destroy()

def obtener_datos():
    # Obtener los datos de los archivos TXT
    try:
        # Pedir fecha
        def procesar_con_fecha(mes, dia, anio):
            fecha_seleccionada = datetime(anio, MESES_NUM[mes], dia)
            # Buscar archivos de tornos
            archivo_torno1, archivo_torno2 = buscar_archivos_torno(fecha_seleccionada)
            if not archivo_torno1 or not archivo_torno2:
                messagebox.showerror("Error", "No se encontraron los archivos de reporte para la fecha seleccionada")
                return
            
            # Leer los datos
            datos_torno1 = leer_archivo_torno(archivo_torno1)
            datos_torno2 = leer_archivo_torno(archivo_torno2)
            if not datos_torno1 or not datos_torno2:
                messagebox.showerror("Error", "No se pudieron leer los archivos de reporte")
                return
            
            # Procesar
            procesar_ambos_tornos(datos_torno1, datos_torno2, mes, dia, anio)
        
        # Mostrar el selector de fecha
        pedir_fecha(procesar_con_fecha)
        
    except Exception as e:
        messagebox.showerror("Error", f"Error al obtener datos: {str(e)}")
        escribir_log(f"Error en obtener_datos: {str(e)}", nivel="error")

def continuar_a_fecha(ventana, widget_texto, datos_torno1):
    """Función para manejar el paso a selección de fecha"""
    datos_torno2 = widget_texto.get("1.0", tk.END).strip()
    if not datos_torno2:
        messagebox.showwarning("Advertencia", "Ingresa los datos del Torno 2.")
        return

    ventana.destroy()
    pedir_fecha(lambda m,d,a: procesar_ambos_tornos(datos_torno1, datos_torno2, m, d, a))

def ejecutar(txt, torno, mes, dia, anio, callback_final=None):
    # Función principal para el procesamiento
    try:
        # Registro inicial en el log
        escribir_log(f"Iniciando procesamiento para Torno {torno} - Fecha: {dia}/{mes}/{anio}")
        inicio_barra = 0 if torno == 1 else 50
        barra['value'] = inicio_barra
        ventana_carga.update_idletasks()
        def incrementar_barra(hasta, paso=1):
            """Función mejorada con registro de progreso"""
            nonlocal inicio_barra
            valor_final = inicio_barra + hasta
            escribir_log(f"Progreso Torno {torno}: {barra['value']}% -> {valor_final}%")
            for i in range(barra['value'], valor_final + 1, paso):
                barra['value'] = i
                ventana_carga.update_idletasks()
                time.sleep(0.01)

        # Paso 1: Obtener rendimientos (10%)
        incrementar_barra(10)
        fecha_actual = datetime(anio, MESES_NUM[mes], dia).date()
        rendimiento_log = obtener_rendimientos_de_log(fecha_actual)
        
        if rendimiento_log:
            escribir_log(f"Rendimientos obtenidos - Torno 1: {rendimiento_log.get('torno1', 'N/A')}%, "
                       f"Torno 2: {rendimiento_log.get('torno2', 'N/A')}%")

        # Paso 2: Preparar hoja (30%)
        escribir_log("Preparando hoja del mes...")
        incrementar_barra(20)
        if not preparar_hoja_mes(mes, dia, anio):
            escribir_log("Error al preparar hoja del mes", nivel="error")
            if callback_final:
                callback_final(False)
            return False

        # Paso 3: Procesar datos (70%)
        escribir_log(f"Procesando datos del Torno {torno}...")
        incrementar_barra(40)
        bloques, porcentajes = procesar_datos(txt, torno, mes, dia, anio)
        if bloques is None or porcentajes is None:
            escribir_log("Error al procesar datos", nivel="error")
            if callback_final:
                callback_final(False)
            return False

        # Paso 4: Escribir en hoja (100%)
        incrementar_barra(30)
        resultado = fecha(mes, dia, anio, torno, bloques, porcentajes, 
                         lambda h: incrementar_barra(h), 
                         rendimiento_log if torno == 2 else None)
        if resultado:
            escribir_log(f"Procesamiento del Torno {torno} completado con éxito \n")
        else:
            escribir_log(f"Error en el procesamiento del Torno {torno}", nivel="error")
        if callback_final and torno == 2:
            callback_final(resultado)
        return resultado

    except Exception as e:
        escribir_log(f"Error crítico en Torno {torno}: {str(e)}", nivel="error")
        if callback_final and torno == 2:
            callback_final(False)
        return False

def obtener_rendimientos_de_log(fecha_ingresada):
    # Obtener los datos de rendimiento del archivo tornos.log
    escribir_log(f"Buscando rendimientos para fecha: {fecha_ingresada}")
    log_path = os.path.join(BASE_DIR, "tornos.log")
    fecha_str = fecha_ingresada.strftime("%Y-%m-%d")
    rendimientos = {'torno1': None, 'torno2': None}

    if not os.path.exists(log_path):
        escribir_log(f"Archivo de log no encontrado: {log_path}", nivel="warning")
        return None

    try:
        with open(log_path, 'r', encoding='utf-8') as f:
            lineas = f.readlines()

        # Patrón para obtener la fecha y los rendimientos
        patron = re.compile(
            r"Fecha:\s*" + re.escape(fecha_str) + 
            r".*?Torno\s*1:\s*Rendimiento:\s*(\d+\.\d+).*?" +
            r"Torno\s*2:\s*Rendimiento:\s*(\d+\.\d+)",
            re.DOTALL | re.IGNORECASE
        )

        # Buscar desde el final hacia atrás para encontrar el dato el más reciente
        for linea in reversed(lineas):
            if f"Fecha: {fecha_str}" in linea:
                bloque = linea
                idx = lineas.index(linea)
                for siguiente in lineas[idx+1:idx+3]:
                    if "Torno" in siguiente:
                        bloque += siguiente
                    else:
                        break
                
                coincidencia = patron.search(bloque)
                if coincidencia:
                    rendimientos['torno1'] = float(coincidencia.group(1))
                    rendimientos['torno2'] = float(coincidencia.group(2))
                    escribir_log(f"Rendimientos encontrados para {fecha_str}: "
                               f"Torno1={rendimientos['torno1']}%, "
                               f"Torno2={rendimientos['torno2']}%")
                    return rendimientos

        escribir_log(f"No se encontraron rendimientos para {fecha_str}", nivel="warning")
        return None

    except Exception as e:
        escribir_log(f"Error al leer el archivo de log: {str(e)}", nivel="error")
        return None

def procesar_ambos_tornos(datos_torno1, datos_torno2, mes, dia, anio):
    # Procesar el torno 1 y luego el 2 y mostar ventana de exito o fallo
    mostrar_carga()
    # Crear copia de seguridad
    try:
        backup_path = os.path.join(BASE_DIR, CARPETA, "Reporte IR Tornos copia_de_seguridad.xlsx")
        shutil.copy(RUTA_ENTRADA, backup_path)
        escribir_log(f"Copia de seguridad creada")
    except Exception as e:
        escribir_log(f"No se pudo crear copia de seguridad inicial: {str(e)}", nivel="error")
        messagebox.showerror("Error", "No se pudo crear la copia de seguridad inicial. Verifique permisos.")
        ventana_carga.destroy()
        return

    def mostrar_resultado_final(exito):
        """Muestra el mensaje final asegurando visibilidad"""
        # Cerrar ventana de carga
        ventana_carga.destroy()
        # Forzar enfoque en la ventana principal
        ventana.attributes('-topmost', True)
        ventana.lift()
        ventana.focus_force()
        ventana.update_idletasks()
        if exito:
            mensaje = tk.Toplevel(ventana)
            mensaje.title("Proceso Completado")
            mensaje.geometry("400x150")
            mensaje.resizable(False, False)
            # Mensaje de exito
            tk.Label(mensaje, 
                    text="Éxito ✅ Valores actualizados correctamente para el día:\n"
                         f"Fecha: {dia}/{mes}/{anio}", pady=20).pack()
            tk.Button(mensaje, text="Aceptar", command=ventana.destroy,
                    width=15).pack(pady=10)
            mensaje.grab_set()
            mensaje.focus_force()
            mensaje.attributes('-topmost', True)
            mensaje.update_idletasks()
            x = ventana.winfo_x() + (ventana.winfo_width() - mensaje.winfo_width()) // 2
            y = ventana.winfo_y() + (ventana.winfo_height() - mensaje.winfo_height()) // 2
            mensaje.geometry(f"+{x}+{y}")
            ventana.attributes('-topmost', False)
        else:
            # Ventana para errores
            ventana.bell()
            messagebox.showerror(
                "Error", 
                "❌ Ocurrió un error durante el procesamiento\n"
                "Revise el archivo de log para más detalles",
                parent=ventana
            )
            ventana.destroy()

    def tarea_principal():
        try:
            # Procesar Torno 1
            if not ejecutar(datos_torno1, 1, mes, dia, anio):
                ventana.after(0, lambda: mostrar_resultado_final(False))
                return

            # Procesar Torno 2
            ejecutar(datos_torno2, 2, mes, dia, anio, 
                    lambda exito: ventana.after(0, lambda: mostrar_resultado_final(exito)))
        except Exception as e:
            escribir_log(f"Error inesperado: {str(e)}", nivel="error")
            ventana.after(0, lambda: mostrar_resultado_final(False))

    threading.Thread(target=tarea_principal, daemon=True).start()

def procesar_datos(entrada, torno, mes, dia, anio):
    # Procesar los datos y escribirlos en el Excel
    escribir_log(f"Inicio de procesar_datos - Torno: {torno}, Fecha: {dia}/{mes}/{anio}")
    bloques_detectados = []
    sumas_ad_por_bloque = []
    # 1. Verificación inicial del archivo
    if not os.path.exists(RUTA_ENTRADA):
        error_msg = f"No se encontró el archivo Excel en:\n{RUTA_ENTRADA}"
        messagebox.showerror("Error", error_msg)
        escribir_log("ERROR - Archivo no encontrado", nivel="error")
        return None, None
    # 2. Verificar de permisos de escritura
    try:
        # Intento de apertura
        with open(RUTA_ENTRADA, 'a+b') as test_file:
            pass
    except PermissionError:
        error_msg = f"El archivo está abierto en Excel. Por favor cierre:\n{RUTA_ENTRADA}"
        messagebox.showerror("Error", error_msg)
        escribir_log("ERROR - El archivo esta abierto \n", nivel="error")
        return None, None
    except Exception as e:
        error_msg = f"No se puede acceder al archivo:\n{str(e)}"
        messagebox.showerror("Error", error_msg)
        escribir_log(f"ERROR - Acceso al archivo: {str(e)}", nivel="error")
        return None, None
    # 3. Procesamiento principal
    wb = None
    try:
        # Intentar abrir el workbook
        wb = openpyxl.load_workbook(RUTA_ENTRADA)
        # Verificar si existe la hoja "IR diario "
        if "IR diario " not in wb.sheetnames:
            error_msg = 'No se encontró la hoja "IR diario " en el archivo Excel'
            messagebox.showerror("Error", error_msg)
            escribir_log("ERROR - Hoja 'IR diario ' no encontrada", nivel="error")
            return None, None
        hoja = wb["IR diario "]
        ultima_fila = None
        # Buscar última fila con patrón "* * ..."
        for fila in hoja.iter_rows():
            if [str(c.value).strip() if c.value else "" for c in fila[:3]] == ["*", "*", "..."]:
                ultima_fila = fila[0].row
        if not ultima_fila:
            raise ValueError("No se encontró '* * ...'")
        fila = ultima_fila + 1
        for b in extraer_bloques(entrada):
            try:
                f_ini = fila
                subs = sub_bloques(b)
                filas_validas = []
                # Procesar cada subbloque
                for sub in subs:
                    txt = sub[0] if not re.match(r'^\d', sub[0]) else ""
                    datos = sub[1:] if txt else sub
                    # Construir datos de columnas
                    p = txt.split()
                    col_txt = (
                        [p[0], p[1], p[2], p[3], "", p[4]] if "*" in txt and len(p) >= 5 and p[0] == "*" else
                        ["*", "*", "...", "", "", ""] if "*" in txt else
                        [p[0], p[1], p[2], p[3], "", p[4]] if len(p) >= 5 else
                        ["", p[0], p[1], p[2], "", p[3]] if len(p) == 4 else
                        [""] * 6
                    )
                    col_nums = [val for l in datos for val in l.strip().split()]
                    fila_vals = col_txt + col_nums
                    # Escribir valores en las celdas (columnas 1-24)
                    for col, val in enumerate(fila_vals[:24], 1):
                        try:
                            n = float(val.replace(",", ".")) if 3 <= col <= 24 and val else val
                            escribir(hoja, fila, col, n, isinstance(n, float))
                        except:
                            escribir(hoja, fila, col, val)
                    # Escribir metadatos (columnas 25-28)
                    for col, val in zip(range(25, 29), [torno, mes, dia, anio]):
                        hoja.cell(row=fila, column=col, value=val).alignment = ALIGN_R
                    fila += 1
                f_fin = fila - 1
                tipo_bloque = "PODADO" if "PODADO" in txt.upper() else "REGULAR"
                bloques_detectados.append((tipo_bloque, f_fin))
                # Insertar fórmulas proporcionales (columna AD)
                if len(subs) > 1:
                    for f in range(f_ini, f_fin):
                        hoja.cell(row=f, column=30, value=f"=IFERROR(AC{f}*D{f}/D{f_fin}, 0)")
                # Configurar celda de autosuma
                for col in range(25, 30):
                    hoja.cell(row=f_fin, column=col, value="")
                celda_autosuma = hoja.cell(row=f_fin, column=30)
                celda_autosuma.value = f"=SUM(AD{f_ini}:AD{f_fin-1})"
                celda_autosuma.fill = FILL_AMARILLO
                bloque_texto = " ".join(b).upper()
                tipo_bloque = "PODADO" if "PODADO" in bloque_texto else "REGULAR"
                valor_d = hoja.cell(row=f_fin, column=4).value
                try:
                    valor_d = float(str(valor_d).replace(",", ".")) if valor_d else 0
                except:
                    valor_d = 0
                bloques_detectados.append((tipo_bloque, valor_d))
                
                if tipo_bloque != "PODADO":
                    for col in range(25, 30):
                        hoja.cell(row=f_fin, column=col, value="")
                try:
                    valor_ae = Pasar_referencia(f"AD{f_fin}")
                    sumas_ad_por_bloque.append(valor_ae)
                except Exception as e:
                    sumas_ad_por_bloque.append(0.0)
                    escribir_log(f"Error al obtener referencia AD{f_fin}: {str(e)}", nivel="warning")
                # Guardar cambios después de cada bloque
                try:
                    wb.save(RUTA_ENTRADA)
                except PermissionError:
                    error_msg = "El archivo Excel fue bloqueado durante la ejecución. Por favor ciérrelo."
                    messagebox.showerror("Error", error_msg)
                    escribir_log("ERROR - Archivo bloqueado durante la ejecución", nivel="error")
                    return None, None
            except Exception as e:
                escribir_log(f"Error procesando bloque: {str(e)}", nivel="error")
                continue
        # En caso de errores
        return bloques_detectados, sumas_ad_por_bloque
    except PermissionError:
        error_msg = "El archivo Excel fue abierto durante la ejecución. Operación cancelada."
        messagebox.showerror("Error", error_msg)
        escribir_log("ERROR - Archivo abierto durante la ejecución", nivel="error")
        return None, None
    except Exception as e:
        error_msg = f"Error inesperado: {str(e)}"
        messagebox.showerror("Error", error_msg)
        escribir_log(f"ERROR - {str(e)}", nivel="error")
        return None, None
    finally:
        if wb is not None:
            try:
                wb.close()
            except:
                pass
        escribir_log("Procesamiento de datos completado")

def escribir(hoja, fila, col, valor, es_numero=False):
    """Escribe un valor en la celda con formato adecuado"""
    celda = hoja.cell(row=fila, column=col, value=valor)
    celda.border = BORDER
    celda.alignment = ALIGN_R
    if es_numero:
        celda.number_format = '0.00'

def Pasar_referencia(celda_origen):
    # Retorna la referencia de la celda requerida
    escribir_log("Inicio de pasar_referencia")
    if not re.match(r'^AD\d+$', celda_origen):
        messagebox.showerror("Error", f"Formato de celda inválido: {celda_origen}")
        escribir_log("Error", f"Formato de celda inválido: {celda_origen}")
        raise ValueError(f"Formato de celda inválido: {celda_origen}")
    referencia = f"='IR diario '!{celda_origen}"
    return referencia

def extraer_bloques(txt):
    escribir_log("Inicio de extraer_bloques")
    lineas = [l.strip() for l in txt.strip().split("\n") if l.strip()]
    bloques, b, i = [], [], 0
    while i < len(lineas):
        l = lineas[i]
        if re.match(r'^\* \* \.\.\.', l):
            b.append(l)
            if i+1 < len(lineas) and re.match(r'^\d', lineas[i+1]):
                b.append(lineas[i+1])
                i += 1
            bloques.append(b)
            b = []
        else: b.append(l)
        i += 1
    if b: bloques.append(b)
    return bloques

def sub_bloques(b):
    escribir_log("Inicio de sub_bloques")
    subs, tmp = [], []
    for l in b:
        if re.match(r'^\D', l) or '*' in l:
            if tmp: subs.append(tmp)
            tmp = [l]
        else: tmp.append(l)
    if tmp: subs.append(tmp)
    return subs

def escribir_valor_bloque(hoja, col_dia, torno, valor, tipo_bloque):
    escribir_log("Inicio de escribir_valor_bloque")
    tipo_bloque = tipo_bloque.strip().upper()
    if tipo_bloque == "PODADO":
        fila_valor = 3 if torno == 1 else 4
    elif tipo_bloque == "REGULAR":
        fila_valor = 8 if torno == 1 else 9
    else:
        messagebox.showwarning("Advertencia", f"Tipo de bloque no reconocido: '{tipo_bloque}'")
        escribir_log("Advertencia", f"Tipo de bloque no reconocido: '{tipo_bloque}'")
        return
    try:
        if valor is None:
            valor_final = 0.0
        elif isinstance(valor, (int, float)):
            valor_final = float(valor)
        else:
            valor_final = float(str(valor).replace(",", "."))
    except ValueError:
        valor_final = 0.0
    celda = hoja.cell(row=fila_valor, column=col_dia)
    celda.value = valor_final
    celda.number_format = '0'

def escribir_valores_resumen_bloques(hoja, col_dia, torno, valores_ae_por_bloque, tipos_bloque, rendimiento_log=None):
    # Escribe referencias y fórmulas en las celdas 
    escribir_log("Inicio de escribir_valores_resumen_bloques")
    letra_actual = openpyxl.utils.get_column_letter(col_dia)
    try:
        # 1. Escribir fórmulas
        # hoja.cell(row=34, column=col_dia, 
        #          value=f"=IFERROR(AVERAGE({letra_actual}32:{letra_actual}33), 0)").number_format = '0.00%'
        hoja.cell(row=38, column=col_dia, 
                 value=f"=IFERROR({letra_actual}32/{letra_actual}23, 0)").font = Font(color='000000')
        hoja.cell(row=39, column=col_dia, 
                 value=f"=IFERROR({letra_actual}33/{letra_actual}24, 0)").font = Font(color='000000')
        hoja.cell(row=40, column=col_dia, 
                 value=f"=IFERROR({letra_actual}34/{letra_actual}28, 0)").number_format = '0.00%'
        hoja.cell(row=40, column=col_dia).font = Font(color='000000')
        # 2. Escribir las referencias de los bloques
        for i, (tipo_bloque, referencia) in enumerate(zip(tipos_bloque, valores_ae_por_bloque)):
            tipo_bloque = tipo_bloque.strip().upper()
            if tipo_bloque == "PODADO":
                fila_valor = 13 if torno == 1 else 14
            elif tipo_bloque == "REGULAR":
                fila_valor = 18 if torno == 1 else 19
            else:
                continue
            celda = hoja.cell(row=fila_valor, column=col_dia)
            celda.value = referencia
            celda.alignment = ALIGN_R
            escribir_log(f"Bloque {i} ({tipo_bloque}) | Torno {torno} | Fila {fila_valor}")
            escribir_log(f"Referencia escrita: {referencia}")
        # Escribir rendimientos del log si existen
        if rendimiento_log and isinstance(rendimiento_log, dict):
            try:
                try:
                    hoja.protection.disable()
                except:
                    pass

                torno1 = float(rendimiento_log.get('torno1', 0))
                torno2 = float(rendimiento_log.get('torno2', 0))
                hoja[f'{letra_actual}32'] = torno1/100
                hoja[f'{letra_actual}33'] = torno2/100
                hoja[f'{letra_actual}32'].number_format = '0.00%'
                hoja[f'{letra_actual}33'].number_format = '0.00%'
                hoja.parent.save(RUTA_ENTRADA)
                escribir_log(f"¡VALORES ESCRITOS! {letra_actual}32: {torno1}%, {letra_actual}33: {torno2}%")

            except Exception as e:
                escribir_log(f"Error al escribir rendimientos: {str(e)}", nivel="error")
                raise

    except Exception as e:
        escribir_log(f"Error crítico: {str(e)}", nivel="error")
        raise

def fecha(mes, dia, anio, torno, bloques_detectados, sumas_ad_por_bloque, incrementar_barra, rendimiento_log=None):
    escribir_log("Inicio de fecha")
    # Escribe los datos en la hoja del mes fechas, datos y las referencias
    nombre_hoja = f"IR {mes} {anio}".strip()
    col_dia = dia + 1
    wb = None
    exito = False
    try:
        wb = openpyxl.load_workbook(RUTA_ENTRADA)
        if nombre_hoja not in wb.sheetnames:
            messagebox.showerror("Error", f"No se encontró la hoja '{nombre_hoja}'")
            escribir_log(f"Error - Hoja '{nombre_hoja}' no encontrada", nivel="error")
            return False
        hoja_mes = wb[nombre_hoja]
        nueva_fecha = f"{dia:02d}/{MESES_NUM[mes]:02d}/{anio}"
        # 2. Escribir la fecha en las celdas correspondientes
        filas_fecha = [2, 7, 12, 17, 22, 27, 31, 37]
        for fila in filas_fecha:
            try:
                celda = hoja_mes.cell(row=fila, column=col_dia)
                celda.value = nueva_fecha
            except Exception as e:
                messagebox.showwarning("Advertencia", f"Error escribiendo fecha en {fila},{col_dia}: {str(e)}")
                escribir_log("Advertencia", f"Error escribiendo fecha en {fila},{col_dia}: {str(e)}")
        # 3. Escribir valores de bloques
        valores_para_escribir = [val for i, (tipo, val) in enumerate(bloques_detectados) if i % 2 == 1]
        tipos_para_escribir = [tipo for i, (tipo, val) in enumerate(bloques_detectados) if i % 2 == 1]
        for (tipo_bloque, valor), valor_ae in zip(zip(tipos_para_escribir, valores_para_escribir), sumas_ad_por_bloque):
            escribir_valor_bloque(hoja_mes, col_dia, torno, valor, tipo_bloque)
            escribir_valores_resumen_bloques(hoja_mes, col_dia, torno, [valor_ae], [tipo_bloque], rendimiento_log)
        # 4. Guardar cambios y crear copia
        shutil.copy(RUTA_ENTRADA, os.path.join(BASE_DIR, ARCHIVO))
        wb.save(RUTA_ENTRADA)
        # 5. Copia de seguridad
        try:
            shutil.copy(RUTA_ENTRADA, os.path.join(BASE_DIR, ARCHIVO))
        except Exception as e:
            messagebox.showwarning("Advertencia", f"No se pudo crear copia de seguridad:\n{str(e)}")
        exito = True
        return exito
    except Exception as e:
        messagebox.showerror("Error", 
            f"No se pudo escribir en hoja:\n{str(e)}\n\n"
            "Verifique que el archivo no esté abierto\n"
                )
        error_msg = (
            f"El documento esta abierto cierrelo e intentelo nuevamente \n"
                )
        escribir_log(error_msg, "error")
        return False
    finally:
        # 6. Cerrar el workbook si está abierto
        if wb is not None:
            try:
                wb.close()
            except:
                messagebox.showwarning("Advertencia", "Error al cerrar el workbook")
                escribir_log("Advertencia", "Error al cerrar el workbook")

        # 7. escribir en el log el mensaje de éxito solo si todo salió bien en cada torno
        if exito:
            escribir_log(f"Éxito ✅ Valores actualizados correctamente")
            escribir_log(f"Fin de la ejecucucion")

# codigo con errores
# def preparar_hoja_mes(mes, dia, anio):
#     import win32com.client as win32, pythoncom
#     import re
#     escribir_log("Inicio de preparar_hoja_mes")
#     nombre_hoja = f"IR {mes} {anio}"

#     try:
#         # Paso 1: Verificar con openpyxl si ya existe
#         wb_check = openpyxl.load_workbook(RUTA_ENTRADA)
#         if nombre_hoja in wb_check.sheetnames:
#             escribir_log(f"La hoja '{nombre_hoja}' ya existe. No se creará una nueva.")
#             wb_check.close()
#             return True
#         wb_check.close()

#         # Paso 2: Crear hoja con win32com
#         pythoncom.CoInitialize()
#         excel = win32.DispatchEx("Excel.Application")
#         excel.Visible = False
#         excel.DisplayAlerts = False
#         wb = excel.Workbooks.Open(os.path.abspath(RUTA_ENTRADA), UpdateLinks=0)

#         # Validar con win32com si ya existe la hoja (verificación fuerte)
#         hoja_existente = None
#         for hoja in wb.Sheets:
#             if hoja.Name == nombre_hoja:
#                 hoja_existente = hoja
#                 break

#         if hoja_existente is not None:
#             escribir_log(f"La hoja '{nombre_hoja}' ya existe. No se creará una nueva.")
#             wb.Close(SaveChanges=False)
#             excel.Quit()
#             pythoncom.CoUninitialize()
#             return True

#         # Buscar hoja anterior válida para copiar
#         hojas = [h.Name for h in wb.Sheets]
#         hojas_ir = [h for h in hojas if re.match(r"^IR\s+\w+\s+\d{4}$", h)]

#         def total_meses(nombre):
#             try:
#                 _, mes_str, anio_str = nombre.split()
#                 return int(anio_str) * 12 + MESES_NUM[mes_str]
#             except:
#                 return -1

#         hojas_ordenadas = sorted(hojas_ir, key=total_meses)
#         total_nueva = int(anio) * 12 + MESES_NUM[mes]
#         hoja_anterior = None

#         for h in hojas_ordenadas:
#             if total_meses(h) < total_nueva:
#                 hoja_anterior = h
#             else:
#                 break

#         if not hoja_anterior:
#             wb.Close(SaveChanges=False)
#             excel.Quit()
#             pythoncom.CoUninitialize()
#             messagebox.showwarning("Error", f"No hay hoja anterior para copiar en {nombre_hoja}")
#             return False

#         # Copiar hoja anterior al final
#         wb.Sheets(hoja_anterior).Copy(After=wb.Sheets(wb.Sheets.Count))
#         nueva_hoja = wb.ActiveSheet
#         nueva_hoja.Name = nombre_hoja

#         escribir_log(f"Hoja '{nombre_hoja}' creada copiando desde '{hoja_anterior}'")

#         # Ajustar gráficos
#         chart_objects = nueva_hoja.ChartObjects()
#         if chart_objects.Count > 0:
#             chart1 = chart_objects(1).Chart
#             chart1.HasTitle = True
#             chart1.ChartTitle.Text = f"IR Diario Tornos {mes} {anio}"
#             chart1.ChartTitle.Font.Size = 12
#             chart1.ChartTitle.Font.Bold = True
#             chart1.Axes(1).HasTitle = True
#             chart1.Axes(1).AxisTitle.Text = " "
#             chart1.Axes(2).HasTitle = True
#             chart1.Axes(2).AxisTitle.Text = " "

#         if chart_objects.Count > 1:
#             chart2 = chart_objects(2).Chart
#             chart2.HasTitle = True
#             chart2.ChartTitle.Text = f"IR v/s R% {mes} {anio}"
#             chart2.ChartTitle.Font.Size = 12
#             chart2.ChartTitle.Font.Bold = True
#             chart2.Axes(1).HasTitle = True
#             chart2.Axes(1).AxisTitle.Text = " "
#             chart2.Axes(2).HasTitle = True
#             chart2.Axes(2).AxisTitle.Text = " "

#         wb.Close(SaveChanges=True)
#         excel.Quit()
#         pythoncom.CoUninitialize()
#         return True

#     except Exception as e:
#         escribir_log(f"Error en preparar_hoja_mes: {str(e)}", nivel="error")
#         messagebox.showerror("Error crítico", f"No se pudo preparar la hoja del mes:\n{str(e)}")
#         return False



#         # Paso 3: Configurar fórmulas y limpieza
#         wb_openpyxl = openpyxl.load_workbook(RUTA_ENTRADA)
#         try:
#             hoja = wb_openpyxl[nombre_hoja]

#             # Limpiar celdas específicas
#             filas_a_limpiar = [2,3,4,7,8,9,12,13,14,17,18,19,22,23,24,27,28,31,32,33,34,37,38,39,40]
#             for fila in filas_a_limpiar:
#                 for col in range(2, 35):
#                     celda = hoja.cell(row=fila, column=col)
#                     if not isinstance(celda, openpyxl.cell.cell.MergedCell):
#                         celda.value = ""

#             # Configurar fórmulas para cada día del mes
#             dias_mes = dias_en_mes(mes, anio)
#             for col in range(2, 2 + dias_mes):
#                 dia_mes = col - 1
#                 fecha = f"{dia_mes:02d}/{MESES_NUM[mes]:02d}/{anio}"
#                 for fila in [2,7,12,17,22,27,31,37]:  # Filas con fechas
#                     hoja.cell(row=fila, column=col, value=fecha)

#                 letra = openpyxl.utils.get_column_letter(col)
#                 # Fórmulas clave
#                 hoja.cell(row=23, column=col, value=f"=IFERROR(({letra}3*{letra}13+{letra}8*{letra}18)/({letra}3+{letra}8), 0)")
#                 hoja.cell(row=24, column=col, value=f"=IFERROR(({letra}4*{letra}14+{letra}9*{letra}19)/({letra}4+{letra}9), 0)")
#                 hoja.cell(row=28, column=col, value=f"=IFERROR(({letra}23*({letra}3+{letra}8)+{letra}24*({letra}4+{letra}9))/({letra}3+{letra}4+{letra}8+{letra}9), 0)")

#             # Configurar resumen mensual (columna AH)
#             hoja.cell(row=2, column=34, value=int(anio))
#             for fila in [3,4,8,9]:
#                 hoja.cell(row=fila, column=34, value=f"=SUM(B{fila}:AG{fila})")
#             hoja.cell(row=23, column=34, value="=IFERROR((AH3*AH13+AH8*AH18)/(AH3+AH8), 0)")
#             hoja.cell(row=24, column=34, value="=IFERROR((AH4*AH14+AH9*AH19)/(AH4+AH9), 0)")
#             hoja.cell(row=28, column=34, value="=IFERROR((AH23*(AH3+AH8)+AH24*(AH4+AH9))/(AH3+AH4+AH8+AH9), 0)")
            
#             # Formato especial
#             hoja.cell(row=32, column=34, value="R%").font = Font(bold=True)
#             hoja.cell(row=38, column=34, value="IR%").font = Font(bold=True)
#             hoja.cell(row=32, column=34).alignment = Alignment(horizontal='center', vertical='center')
#             hoja.cell(row=38, column=34).alignment = Alignment(horizontal='center', vertical='center')
            
#             # Porcentajes
#             hoja.cell(row=39, column=34, value="=AH33/AH28").number_format = '0.00%'
#             hoja.cell(row=40, column=34, value="=AH34/AH28").number_format = '0.00%'

#             # Limpiar celdas adicionales
#             for fila in [49,50,51]:
#                 for col_limpiar in [27,28,29,30]:
#                     hoja.cell(row=fila, column=col_limpiar, value=" ")

#             wb_openpyxl.save(RUTA_ENTRADA)
#         finally:
#             wb_openpyxl.close()

#         pythoncom.CoInitialize()
#         excel = win32.DispatchEx("Excel.Application")
#         excel.Visible = False
#         excel.DisplayAlerts = False
        
#         try:
#             wb_calc = excel.Workbooks.Open(os.path.abspath(RUTA_ENTRADA), UpdateLinks=0)
#             excel.CalculateFull()
#             wb_calc.Save()
#         finally:
#             wb_calc.Close()
#             excel.Quit()
#             pythoncom.CoUninitialize()

#         escribir_log(f"Hoja {nombre_hoja} creada exitosamente")
#         return True

#     except Exception as e:
#         error_msg = f"Error en preparar_hoja_mes: {str(e)}"
#         escribir_log(error_msg)
#         messagebox.showerror("Error crítico", error_msg)
#         return False


# ------------------------------------------------------------


# def preparar_hoja_mes(mes, dia, anio):
#     # Crea la hoja del mes si no existe y la configura con fórmulas
#     escribir_log("Inicio de preparar_hoja_mes")
#     nombre_hoja = f"IR {mes} {anio}"
#     col_dia = dia + 1
#     try:
#         # Paso 1: Verificar si la hoja ya existe
#         wb_check = openpyxl.load_workbook(RUTA_ENTRADA)
#         if nombre_hoja in wb_check.sheetnames:
#             hoja_existente = wb_check[nombre_hoja]
#             celdas_clave = [
#                 hoja_existente.cell(row=3, column=col_dia).value,
#                 hoja_existente.cell(row=4, column=col_dia).value,
#                 hoja_existente.cell(row=8, column=col_dia).value,
#                 hoja_existente.cell(row=9, column=col_dia).value
#             ]
#             if any(cell is not None and str(cell).strip() != "" for cell in celdas_clave):
#                 escribir_log(f"El día {dia} ya tiene datos en {nombre_hoja}")
#                 wb_check.close()
#                 return True
#             else:
#                 escribir_log(f"La hoja {nombre_hoja} ya existe y se usará tal cual.")
#                 wb_check.close()
#                 return True
#         wb_check.close()
#         # Paso 2: Crear hoja nueva si no existe
#         import win32com.client as win32, pythoncom
#         pythoncom.CoInitialize()
#         excel = win32.DispatchEx("Excel.Application")
#         excel.Visible = False
#         excel.DisplayAlerts = False
#         wb = excel.Workbooks.Open(os.path.abspath(RUTA_ENTRADA), UpdateLinks=0)
#         hojas = [h.Name for h in wb.Sheets]
#         if nombre_hoja not in hojas:
#             # Buscar hoja anterior para copiar
#             hojas_ir = [h for h in hojas if h.startswith("IR ") and len(h.split()) == 3]
#             def total_meses(nombre):
#                 try:
#                     _, mes_str, anio_str = nombre.split()
#                     return int(anio_str) * 12 + MESES_NUM[mes_str]
#                 except:
#                     return -1
#             hojas_ordenadas = sorted(hojas_ir, key=total_meses)
#             total_nueva = int(anio) * 12 + MESES_NUM[mes]
#             hoja_anterior = None
#             for h in hojas_ordenadas:
#                 if total_meses(h) < total_nueva:
#                     hoja_anterior = h
#                 else:
#                     break
#             if not hoja_anterior:
#                 messagebox.showwarning("Orden inválido", f"No se encontró hoja anterior para '{nombre_hoja}'")
#                 wb.Close(SaveChanges=False)
#                 excel.Quit()
#                 pythoncom.CoUninitialize()
#                 return False
#             idx_anterior = hojas.index(hoja_anterior)
#             insert_idx = min(idx_anterior + 2, wb.Sheets.Count)
#             wb.Sheets(hoja_anterior).Copy(After=wb.Sheets(insert_idx - 1))
#             nueva_hoja = wb.ActiveSheet
#             nueva_hoja.Name = nombre_hoja
#             shutil.copy(RUTA_ENTRADA, os.path.join(BASE_DIR, ARCHIVO))
#             wb.Save()
#         wb.Close(SaveChanges=True)
#         excel.Quit()
#         pythoncom.CoUninitialize()
#         # Paso 3: Configurar hoja
#         wb2 = openpyxl.load_workbook(RUTA_ENTRADA)
#         hoja = wb2[nombre_hoja]
#         filas_a_limpiar = [2,3,4,7,8,9,12,13,14,17,18,19,22,23,24,27,28,31,32,33,34,37,38,39,40]
#         for fila in filas_a_limpiar:
#             for col in range(2, 35):
#                 celda = hoja.cell(row=fila, column=col)
#                 if not isinstance(celda, openpyxl.cell.cell.MergedCell):
#                     celda.value = ""
#         dias_mes = dias_en_mes(mes, anio)
#         for col in range(2, 2 + dias_mes):
#             dia_mes = col - 1
#             fecha = f"{dia_mes:02d}/{MESES_NUM[mes]:02d}/{anio}"
#             for fila in [2,7,12,17,22,27,31,37]:
#                 hoja.cell(row=fila, column=col, value=fecha)
#         for col in range(2, 2 + dias_mes):
#             letra = openpyxl.utils.get_column_letter(col)
#             hoja.cell(row=23, column=col, value=f"=IFERROR(({letra}3*{letra}13+{letra}8*{letra}18)/({letra}3+{letra}8), 0)")
#             hoja.cell(row=24, column=col, value=f"=IFERROR(({letra}4*{letra}14+{letra}9*{letra}19)/({letra}4+{letra}9), 0)")
#             hoja.cell(row=28, column=col, value=f"=IFERROR(({letra}23*({letra}3+{letra}8)+{letra}24*({letra}4+{letra}9))/({letra}3+{letra}4+{letra}8+{letra}9), 0)")
#         hoja.cell(row=32, column=34, value="R%").font = Font(bold=True)
#         hoja.cell(row=38, column=34, value="IR%").font = Font(bold=True)
#         hoja.cell(row=32, column=34).alignment = hoja.cell(row=38, column=34).alignment = Alignment(horizontal='center', vertical='center')
#         for fila in [49,50,51]:
#             for col_limpiar in [27,28,29,30]:
#                 hoja.cell(row=fila, column=col_limpiar, value=" ")
#         hoja.cell(row=2, column=34, value=int(anio))
#         for fila in [3,4,8,9]:
#             hoja.cell(row=fila, column=34, value=f"=SUM(B{fila}:AG{fila})")
#         hoja.cell(row=23, column=34, value="=(AH3*AH13+AH8*AH18)/(AH3+AH8)")
#         hoja.cell(row=24, column=34, value="=(AH4*AH14+AH9*AH19)/(AH4+AH9)")
#         hoja.cell(row=28, column=34, value="=(AH23*(AH3+AH8)+AH24*(AH4+AH9))/(AH3+AH4+AH8+AH9)")
#         hoja.cell(row=39, column=34, value="=AH33/AH28").number_format = '0.00%'
#         hoja.cell(row=40, column=34, value="=AH34/AH28").number_format = '0.00%'
#         wb2.save(RUTA_ENTRADA)
#         wb2.close()
#         # Recalcular fórmulas
#         pythoncom.CoInitialize()
#         excel = win32.DispatchEx("Excel.Application")
#         excel.Visible = False
#         excel.DisplayAlerts = False
#         wb_calc = excel.Workbooks.Open(os.path.abspath(RUTA_ENTRADA), UpdateLinks=0)
#         excel.CalculateFull()
#         wb_calc.Save()
#         wb_calc.Close()
#         excel.Quit()
#         pythoncom.CoUninitialize()
#         return True
#     except Exception as e:
#         messagebox.showerror("Error crítico", f"No se pudo completar la operación:\n{str(e)}")
#         return False


def dias_en_mes(mes, anio):
    escribir_log("Inicio de dias_en_mes")
    # Devuelve el número de días en un mes, considerando años bisiestos
    if mes == "Febrero":
        # Año bisiesto si es divisible por 4, pero no por 100, a menos que también sea divisible por 400
        if (anio % 4 == 0 and anio % 100 != 0) or (anio % 400 == 0):
            return 29
        return 28
    meses_31_dias = ["Enero", "Marzo", "Mayo", "Julio", "Agosto", "Octubre", "Diciembre"]
    return 31 if mes in meses_31_dias else 30



# --- NUEVAS FUNCIONES ---
def hoja_existe_y_es_valida(nombre_hoja, dia):
    """Función mejorada para verificar si una hoja existe y tiene datos"""
    try:
        wb = openpyxl.load_workbook(RUTA_ENTRADA)
        if nombre_hoja not in wb.sheetnames:
            wb.close()
            return False
        
        hoja = wb[nombre_hoja]
        # Verificar más celdas y con mayor tolerancia
        celdas_verificar = [
            (3, dia+1), (4, dia+1), (8, dia+1), (9, dia+1),  # Datos específicos
            (2, 2), (7, 2), (12, 2), (17, 2)  # Encabezados
        ]
        
        tiene_datos = any(
            hoja.cell(row=fila, column=col).value is not None 
            and str(hoja.cell(row=fila, column=col).value).strip() != ""
            for fila, col in celdas_verificar
        )
        
        wb.close()
        return tiene_datos
    except Exception as e:
        escribir_log(f"Error al verificar hoja: {str(e)}", nivel="error")
        return True  # Asumir que existe para evitar sobrescritura

def crear_hoja_mes(mes, anio):
    """Versión final que maneja el error de selección"""
    excel = None
    wb = None
    try:
        nombre_hoja = f"IR {mes} {anio}"
        escribir_log(f"Iniciando creación de {nombre_hoja}")

        # 1. Inicialización COM
        pythoncom.CoInitialize()
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        excel.DisplayAlerts = False  # Deshabilitar alertas adicionales

        # 2. Abrir archivo con reintentos
        for intento in range(3):
            try:
                wb = excel.Workbooks.Open(os.path.abspath(RUTA_ENTRADA), UpdateLinks=0, ReadOnly=False)
                break
            except Exception as e:
                if intento == 2:
                    escribir_log(f"Error abriendo archivo: {str(e)}", nivel="error")
                    return False
                time.sleep(3)

        # 3. Verificar si la hoja ya existe
        try:
            hojas_existentes = [sheet.Name for sheet in wb.Sheets]
            if nombre_hoja in hojas_existentes:
                escribir_log(f"Hoja {nombre_hoja} ya existe")
                wb.Close(SaveChanges=False)
                excel.Quit()
                pythoncom.CoUninitialize()
                return True
        except Exception as e:
            escribir_log(f"Error verificando hojas: {str(e)}", nivel="error")
            return False

        # 4. Encontrar hoja anterior más reciente (evitando 'IR diario')
        def obtener_fecha(nombre):
            try:
                partes = nombre.split()
                if len(partes) == 3 and partes[0] == "IR" and partes[1] in MESES_NUM:
                    return (int(partes[2]), MESES_NUM[partes[1]])
                return (0, 0)
            except:
                return (0, 0)

        hojas_validas = [s for s in hojas_existentes if s.startswith("IR ") and s != nombre_hoja and not s.endswith("diario ")]
        if not hojas_validas:
            escribir_log("No hay hojas válidas para copiar", nivel="error")
            wb.Close(SaveChanges=False)
            excel.Quit()
            pythoncom.CoUninitialize()
            return False

        hoja_origen = max(hojas_validas, key=obtener_fecha)
        escribir_log(f"Copiando desde {hoja_origen}")

        # 5. Método de copiado ultra-robusto
        try:
            # Intentar copiar sin seleccionar (evita el error)
            origen = wb.Sheets(hoja_origen)
            
            # Método 1: Usar Copy directamente
            try:
                origen.Copy(After=wb.Sheets(wb.Sheets.Count))
                time.sleep(5)  # Espera extendida
            except Exception as e1:
                escribir_log(f"Intento 1 falló: {str(e1)}", nivel="warning")
                # Método 2: Alternativa usando API diferente
                try:
                    wb.Sheets.Add(After=wb.Sheets(wb.Sheets.Count)).Name = "TEMP_COPY"
                    temp_sheet = wb.Sheets("TEMP_COPY")
                    origen.Cells.Copy(temp_sheet.Cells)
                    time.sleep(3)
                except Exception as e2:
                    escribir_log(f"Intento 2 falló: {str(e2)}", nivel="error")
                    raise Exception(f"Todos los métodos fallaron: {str(e1)} | {str(e2)}")

            # Verificar nueva hoja
            nueva_hoja = None
            for i in range(wb.Sheets.Count, 0, -1):
                if wb.Sheets(i).Name not in hojas_existentes + ["TEMP_COPY"]:
                    nueva_hoja = wb.Sheets(i)
                    break

            if not nueva_hoja:
                raise Exception("No se detectó nueva hoja creada")

            # Renombrar
            nueva_hoja.Name = nombre_hoja
            
            # Eliminar hoja temporal si existe
            if "TEMP_COPY" in [s.Name for s in wb.Sheets]:
                wb.Sheets("TEMP_COPY").Delete()
            
            # Verificación final
            if nombre_hoja not in [s.Name for s in wb.Sheets]:
                raise Exception("Verificación post-copiado falló")
            
            # Guardar cambios
            wb.Save()
            escribir_log(f"Hoja {nombre_hoja} creada exitosamente")
            return True
            
        except Exception as e:
            escribir_log(f"Error crítico en copiado: {str(e)}", nivel="error")
            # Limpieza de emergencia
            try:
                if nombre_hoja in [s.Name for s in wb.Sheets]:
                    wb.Sheets(nombre_hoja).Delete()
                if "TEMP_COPY" in [s.Name for s in wb.Sheets]:
                    wb.Sheets("TEMP_COPY").Delete()
                wb.Save()
            except:
                pass
            return False
            
    except Exception as e:
        escribir_log(f"Error global: {str(e)}", nivel="error")
        return False
    finally:
        # Limpieza garantizada
        try:
            if wb is not None:
                wb.Close(SaveChanges=True)
        except:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except:
            pass
        try:
            pythoncom.CoUninitialize()
        except:
            pass


def preparar_hoja_mes(mes, dia, anio):
    """Versión simplificada para usar con la nueva función"""
    nombre_hoja = f"IR {mes} {anio}"
    
    # Verificación básica
    try:
        wb = openpyxl.load_workbook(RUTA_ENTRADA)
        if nombre_hoja in wb.sheetnames:
            wb.close()
            return True
        wb.close()
    except Exception as e:
        escribir_log(f"Error en verificación: {str(e)}", nivel="warning")
    
    # Creación de nueva hoja
    if not crear_hoja_mes(mes, anio):
        messagebox.showerror("Error", "No se pudo crear la hoja del mes")
        return False
    
    # Configuración mínima necesaria
    try:
        wb = openpyxl.load_workbook(RUTA_ENTRADA)
        hoja = wb[nombre_hoja]
        
        # Limpieza conservadora
        for fila in [3, 4, 8, 9]:  # Solo filas críticas
            for col in range(2, 32):
                try:
                    celda = hoja.cell(row=fila, column=col)
                    if not isinstance(celda, openpyxl.cell.cell.MergedCell):
                        celda.value = ""
                except:
                    continue
        
        wb.save(RUTA_ENTRADA)
        wb.close()
        return True
        
    except Exception as e:
        escribir_log(f"Error en configuración: {str(e)}", nivel="error")
        try:
            wb.close()
        except:
            pass
        return False

# def preparar_hoja_mes(mes, dia, anio):
#     """Versión mejorada de la función original"""
#     nombre_hoja = f"IR {mes} {anio}"
#     col_dia = dia + 1
    
#     # 1. Verificación robusta de existencia
#     if hoja_existe_y_es_valida(nombre_hoja, dia):
#         escribir_log(f"Usando hoja existente: {nombre_hoja}")
#         return True
    
#     # 2. Creación de nueva hoja
#     if not crear_hoja_mes(mes, anio):
#         messagebox.showerror("Error", "No se pudo crear la hoja para el nuevo mes")
#         return False
    
#     # 3. Configuración inicial de la hoja (manteniendo tu lógica original)
#     try:
#         wb = openpyxl.load_workbook(RUTA_ENTRADA)
#         hoja = wb[nombre_hoja]
        
#         # Limpieza de celdas
#         filas_a_limpiar = [2,3,4,7,8,9,12,13,14,17,18,19,22,23,24,27,28,31,32,33,34,37,38,39,40]
#         for fila in filas_a_limpiar:
#             for col in range(2, 35):
#                 celda = hoja.cell(row=fila, column=col)
#                 if not isinstance(celda, openpyxl.cell.cell.MergedCell):
#                     celda.value = ""
        
#         # Configuración de fechas
#         dias_mes = dias_en_mes(mes, anio)
#         for col in range(2, 2 + dias_mes):
#             dia_mes = col - 1
#             fecha = f"{dia_mes:02d}/{MESES_NUM[mes]:02d}/{anio}"
#             for fila in [2,7,12,17,22,27,31,37]:
#                 hoja.cell(row=fila, column=col, value=fecha)
        
#         # Configuración de fórmulas
#         for col in range(2, 2 + dias_mes):
#             letra = openpyxl.utils.get_column_letter(col)
#             hoja.cell(row=23, column=col, value=f"=IFERROR(({letra}3*{letra}13+{letra}8*{letra}18)/({letra}3+{letra}8), 0)")
#             hoja.cell(row=24, column=col, value=f"=IFERROR(({letra}4*{letra}14+{letra}9*{letra}19)/({letra}4+{letra}9), 0)")
#             hoja.cell(row=28, column=col, value=f"=IFERROR(({letra}23*({letra}3+{letra}8)+{letra}24*({letra}4+{letra}9))/({letra}3+{letra}4+{letra}8+{letra}9), 0)")
        
#         # Configuración de resumen mensual
#         hoja.cell(row=2, column=34, value=int(anio))
#         for fila in [3,4,8,9]:
#             hoja.cell(row=fila, column=34, value=f"=SUM(B{fila}:AG{fila})")
#         hoja.cell(row=23, column=34, value="=(AH3*AH13+AH8*AH18)/(AH3+AH8)")
#         hoja.cell(row=24, column=34, value="=(AH4*AH14+AH9*AH19)/(AH4+AH9)")
#         hoja.cell(row=28, column=34, value="=(AH23*(AH3+AH8)+AH24*(AH4+AH9))/(AH3+AH4+AH8+AH9)")
#         hoja.cell(row=39, column=34, value="=AH33/AH28").number_format = '0.00%'
#         hoja.cell(row=40, column=34, value="=AH34/AH28").number_format = '0.00%'
        
#         wb.save(RUTA_ENTRADA)
#         return True
#     except Exception as e:
#         messagebox.showerror("Error crítico", f"No se pudo completar la operación:\n{str(e)}")
#         return False



if __name__ == "__main__":
    ventana = tk.Tk()
    ventana.title("Procesador de Reportes de Tornos")
    ventana.geometry("400x200")  # Ventana más pequeña y centrada
    tk.Label(ventana, 
             text="Procesador Automático de Reportes",
             font=("Arial", 14)).pack(pady=20)
    
    # Botón de inicio principal
    btn_iniciar = tk.Button(
        ventana,
        text="Iniciar Proceso",
        command=obtener_datos,
        width=20,
        height=2
    )
    btn_iniciar.pack(pady=20)
    
    ventana.mainloop()
